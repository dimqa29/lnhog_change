__author__ = "Dmitry Nikitin"
__version__ = "1.0"

import re
import sys
import time
import os
import json
import collections
from typing import List, Any
import openpyxl
import enetsdk as Enet
from enetsdk.Framework.cm_handler import EMSHandler
from enetconfig.config import ConfigManager
from enetsdk.Utility import ExcelWrapper
from tiermap.tier_mapper import TierMapper
from tiermap import exceptions
import enetsdk

EMAIL_SUBJECT = "Eden-NET SON - "
EMAIL_SUBJECT_ALARM = "Alarm_changes "

USER_PARAMETERS = [
    (
        "Email Addresses (Optional)",
        "Email Addresses to which the agrregated report should be sent. Separate email addresses with semi-colon or comma.",
        Enet.ENET_PARAM_TYPE_STRING,
        "",
        None
    ),
    (
        "Report prefix",
        "Prexif for name of report",
        Enet.ENET_PARAM_TYPE_STRING,
        "",
        None
    ),
    (
        "Audit Type",
        "Audit all target cells or only target cells that are not locked. If \"Unlocked cell only\" is selected, then blocked cells will be ignored.",
        Enet.ENET_PARAM_TYPE_STRING_SET,
        "Unlocked cell only",
        ["All cells", "Unlocked cell only"]
    )
]


class Report_Config(ConfigManager):
    @property
    def emails(self):
        return self.module_parameters.get("Email Addresses (Optional)", "")

    @property
    def Report_prefix(self):
        return self.module_parameters.get("Report prefix", "")

    @property
    def audit_type(self):
        return self.module_parameters.get("Audit Type", "")

    @property
    def excluded_type(self):
        return self.module_parameters.get("Excluded Type", "")


class MyModuleConfig(ConfigManager):

    @property
    def Max_num_changes(self):
        return self.get_int("Max_num_changes_to_push", default=1000)

    @property
    def Email_Alarm(self):
        return self.get_str("Email_alarm_changes", default="")

    @property
    def OSS_value(self):
        return self.get_bool("oss_value", default=True)

    @property
    def amount_in_request(self):
        return self.get_int("amount_in_request", default=300)


    @property
    def check_sw(self):
        return self.get_str("check_sw", default="Yes")

    @property
    def sw_version(self):
        return self.get_str("sw_version", default=["SBTS19A"])




class CheckingChanges:

    @staticmethod
    def _creator_dn(check: (list, dict), etalon: (list, dict)):
        """
        Forming a set from a dictionary or from a list
        :param check: List or dictionary of read values from the network
        :param etalon: List or dictionary of changes sent to the network
        :return: dn_checked and dn_etalon - set consisting of the keys of a dictionary or list
        """
        dn_checked = set(check.keys()) if check and isinstance(check, dict) else set(
            check)
        dn_etalon = set(etalon.keys()) if etalon and isinstance(etalon, dict) else set(
            etalon)

        return dn_checked, dn_etalon

    def check_create_and_update(self, checked_obj: (list, dict), etalon_obj: (list, dict)):
        """
        Checking the correctness of the created elements, set parameters and their values.
        If the DN is not found, such an element is considered not created.
        :param checked_obj: dict
        :param etalon_obj: dict
        :return:
        value_is_set - The value of the parameter on the network corresponds to the requested
            dict: {dn:{param_1: [required value, "OK"], param_2: [required value, "OK"]}, dn_2:...}.
        diff_value - The value of the parameter in the network does not match the requested one
            dict: {dn:{param_1: [required value, real value], param_2: [required value, real value]}, dn_2:...}
        obj_not_create - Dictionary with not found elements in the network
            dict: {dn_1: {param_1: [required value, "does not exist"], param_2: [required value, "N/A"]}}.
        """

        value_is_set = {}
        diff_value = {}  # this variable contains the value of parameters that differ from the reference
        obj_not_create = {}  # this variable contains parameter values whose objects are missing

        set_dn_checked, set_dn_etalon = self._creator_dn(checked_obj, etalon_obj)
        if not set_dn_checked:
            for k in set_dn_etalon:
                par_val_etalon = etalon_obj.get(k)
                obj_not_create[k] = {z: [par_val_etalon[z], "does not exist"] for z in par_val_etalon}
                # not_create_obj[k] = "does not exist"
            return value_is_set, diff_value, obj_not_create

        missing_dn = set_dn_etalon - set_dn_checked
        if len(missing_dn) > 0:
            for k in missing_dn:
                par_val_etalon = etalon_obj.get(k)
                obj_not_create[k] = {z: [par_val_etalon[z], "does not exist"] for z in par_val_etalon}

        crossed_dn = set_dn_checked.intersection(set_dn_etalon)
        if len(crossed_dn) < 1:
            for k in set_dn_etalon:
                par_val_etalon = etalon_obj.get(k)
                obj_not_create[k] = {z: [par_val_etalon[z], "does not exist"] for z in par_val_etalon}
            return value_is_set, diff_value, obj_not_create
        else:
            for i in crossed_dn:
                par_val_check = checked_obj.get(i)
                if par_val_check:
                    set_val_check = set(par_val_check)
                else:
                    par_val_etalon = etalon_obj.get(i)
                    obj_not_create[i] = {z: [par_val_etalon[z], "does not exist"] for z in par_val_etalon}
                    continue

                par_val_etalon = etalon_obj.get(i)
                set_val_etalon = set(par_val_etalon)

                crossed_keys = set_val_check.intersection(set_val_etalon)
                differents = {z: [par_val_etalon[z], par_val_check[z]] for z in crossed_keys if
                              str(par_val_check[z]) != str(par_val_etalon[z])}

                if differents:
                    diff_value[i] = differents

                missing_keys = set_val_etalon - set_val_check
                missing = {z: [par_val_etalon[z], "N/A"] for z in missing_keys if missing_keys}
                if missing:
                    diff_value.setdefault(i, {}).update(missing)

                same = {z: [par_val_etalon[z], "OK"] for z in crossed_keys if
                        str(par_val_check[z]) == str(par_val_etalon[z])}
                if same:
                    value_is_set[i] = same

        return value_is_set, diff_value, obj_not_create

    def check_delete(self, checked_obj, etalon_obj):

        obj_del = {}
        obj_not_del = {}
        set_dn_checked, set_dn_etalon = self._creator_dn(checked_obj, etalon_obj)

        if not set_dn_checked:
            for i in set_dn_etalon:
                obj_del[i] = "delete"
            return obj_del, obj_not_del

        missing = set_dn_etalon - set_dn_checked
        if missing:
            for m in missing:
                obj_del[m] = "delete"

        crossed = set_dn_etalon.intersection(set_dn_checked)
        if crossed:
            for c in crossed:
                obj_not_del[c] = "not delete"

        return obj_del, obj_not_del

class ReadSetData():

    def __init__(self, script_data):
        self.script_data = script_data
        self.ems = self.script_data.GetEmsService()

    def read_date(self, get_params):

        data_preparation = {}
        plug_for_del = ['moClassId',]
        if isinstance(get_params, list):
            for k in get_params:
                data_preparation[k] = plug_for_del
        else:
            for i, o in get_params.items():
                list_params = []
                for a in o:
                    list_params.append(a)
                data_preparation[i] = list_params
        real_val = self.ems.getEMSAttributes(data_preparation)
        return real_val


class ReportGenerator(object):

    def __init__(self, outputfile, start_time, stop_time):
        self.outputfile = outputfile
        self.start_time = start_time
        self.stop_time = stop_time
        self.reporter = ExcelWrapper(self.outputfile)

    def saved_file(self):
        self.reporter.SaveWorkbook()
        return os.path.exists(self.outputfile)

    def quick_builder(self, date):
        self.reporter.QuickBuild(date)

    def create_new_sheets(self, name_sheet, header: list):
        self.reporter.AddWorksheets([name_sheet])
        self.reporter.UseSheet(name_sheet)
        self.reporter.AddHeaders(header, color="FFFF6600", bold=False)
        return self.reporter

    def data_transformation(self, dict_to_modify):

        self.dict_to_modify = dict_to_modify
        date_complited = []
        for key_to, val_to in self.dict_to_modify.items():
            if isinstance(val_to, dict):
                for params_val in val_to.items():
                    d = []
                    d.append(key_to)
                    for i in params_val:
                        if isinstance(i, list):
                            for z in i:
                                if isinstance(z, list):
                                    g = str(z)
                                    d.append(g)
                                else:
                                    d.append(z)
                        else:
                            d.append(i)
                    date_complited.append(d)
            elif isinstance(val_to, list):
                d = []
                d.append(key_to)
                d.extend([', '.join(str(e) for e in val_to)])
                date_complited.append(d)
            else:
                d = []
                d.append(key_to)
                d.append(val_to)
                date_complited.append(d)

        return date_complited


def ScriptMain(script_data, _):
    print("!!start!!")
    start_time = time.strftime("%d-%m-%Y %H:%M:%S")
    time_stamp_start = time.strftime("%d_%m_%Y_%H-%M-%S")
    print(start_time)

    cells = script_data.GetTargets()
    ems_client = EMSHandler(script_data)
    SON_MODE = script_data.GetParameters()['SON Operation Mode']
    config_ini = MyModuleConfig(script_data)
    maximum_changes = config_ini.Max_num_changes
    email_str_alarm = config_ini.Email_Alarm
    oss_value_ini = config_ini.OSS_value
    dns_per_pull_ini = config_ini.amount_in_request
    region_name = script_data.GetRegionName()
    module_name = script_data.GetModuleName()
    config_report = Report_Config(script_data)
    config_data = script_data.GetConfig("Custom_Configuration")
    pathout = script_data.GetUserOutputFilesLoc()
    outputfile = pathout + module_name + "_" + region_name + "_" + config_report.Report_prefix + "_" + time_stamp_start + ".xlsx"
    reporter = ExcelWrapper(outputfile)
    email_str = config_report.emails

    all_cells = len(cells)
    create = {}
    update = {}
    delete = []

    print("*************************************************************************")
    print("Settings from ini file:")
    print("Maximum changes to push: {} ".format(maximum_changes))
    print("Address for Email_Alarm_changes: {} ".format(email_str_alarm))
    print("Read from OSS: {}".format(oss_value_ini))
    print("Number of query: {}".format(dns_per_pull_ini))
    print("*************************************************************************")

    if config_data and isinstance(config_data, dict):
        filename = config_data.get('file_name')
        io_stream = config_data.get('raw config file')
    else:
        print("No active configuration file found. Stop")
        exit()
    wb = openpyxl.load_workbook(io_stream or filename)
    default_param = parser_exel(wb)
    str_v_dict_1 = default_param.get("b2Threshold1GERANQci1")
    b2Threshold1GERANQci1_dict = json.loads(str_v_dict_1)
    str_v_dict_2 = default_param.get("b2Threshold1GERAN")
    b2Threshold1GERAN_dict = json.loads(str_v_dict_2)
    default_param['b2Threshold1GERANQci1'] = b2Threshold1GERANQci1_dict
    default_param['b2Threshold1GERAN'] = b2Threshold1GERAN_dict

    list_dn_cell = []
    dict_dn_cell_list_bcch = {}
    for cell in cells:
        dn_cell = cell.dn
        list_dn_cell.append(dn_cell)
        # sosedi
        iratNeighbors = cell.GetNeighbors(neighborType="irat", ignore_missing_data=True)
        gsm_neighbors = []
        for neighbor in iratNeighbors:
            list_bcchNo_neib = []
            if neighbor.technology == "GSM":
                gsm_neighbors.append(neighbor)
            for gsm_neib in gsm_neighbors:
                bcchNo_neib = gsm_neib.bcch_frequency
                list_bcchNo_neib.append(str(bcchNo_neib))
                list_bcchNo_neib = list(set(list_bcchNo_neib))
                dict_dn_cell_list_bcch[dn_cell+"/LNHOG-0"] = list_bcchNo_neib

    dict_lnhog_param = ems_client.get_child_attributes(list_dn_cell, "LNHOG",
                                                       {"LNHOG": ["arfcnValueListGERAN", "bandIndicatorGERAN", "reportIntervalGERAN",
                                                                  "nccperm", "hysB2ThresholdGERAN", "b2Threshold2RssiGERAN",
                                                                  "b2TimeToTriggerGERANMeas",
                                                                  "b2Threshold1GERANQci1", "b2Threshold1GERAN", "b2Threshold2RssiGERANQci1"]},
                                                       return_by_dn=False)

    dict_get_ems_atribut_FDD = ems_client.get_child_attributes(list_dn_cell, "LNCEL_FDD", {"LNCEL_FDD": ["earfcnDL", "dlChBw"]}, return_by_dn=False)
    dict_get_ems_atribut_TDD = ems_client.get_child_attributes(list_dn_cell, "LNCEL_TDD", {"LNCEL_TDD": ["earfcn", "chBw"]}, return_by_dn=False)
    dict_get_ems_atribut_FDD_TDD = {**dict_get_ems_atribut_FDD, **dict_get_ems_atribut_TDD}
    dict_get_ems_atribut_FDD_TDD_new = {}
    for dn_cel_FDD_TDD, value_FDD_TDD in dict_get_ems_atribut_FDD_TDD.items():
        split_dn_cell = dn_cel_FDD_TDD.split("/", 4)
        dn_cell_LNHOG_FDD_TDD = split_dn_cell[0] + "/" + split_dn_cell[1] + "/" + split_dn_cell[2] + "/" + split_dn_cell[3] + "/LNHOG-0"
        dict_get_ems_atribut_FDD_TDD_new[dn_cell_LNHOG_FDD_TDD] = value_FDD_TDD

    all_dn_lnhog_current = []
    for dn_lnhog_current in dict_lnhog_param.keys():
        all_dn_lnhog_current.append(dn_lnhog_current)


    # update
    # for dn, value in dict_dn_cell_list_bcch.items():
    #     slovar = {}
    #     value_lnhog = dict_lnhog_param.get(dn, {})
    #     for key_1, value_1 in value_lnhog.items():
    #         if sorted(value_1) !=  sorted(value):
    #             slovar['arfcnValueListGERAN'] = value
    #             update[dn] = slovar
    # print("update", update)

    dict_get_ems_atribut_FDD_TDD_new

    for dn, value in dict_dn_cell_list_bcch.items():
        update1 = {}
        slovar = {}
        nccperm_255 = {}
        b2Threshold1GERAN_m120 = {}
        b2Threshold1GERANQci1_m120 = {}
        value_lnhog = dict_lnhog_param.get(dn, {})
        for key_1, value_1 in value_lnhog.items():
            if key_1 == "b2Threshold1GERANQci1" and value_1 != str(default_param.get("b2Threshold1GERANQci1").get(band_bw(dn, dict_get_ems_atribut_FDD_TDD_new), -120)):
                b2Threshold1GERANQci1_m120[key_1] = str(default_param.get("b2Threshold1GERANQci1").get(band_bw(dn, dict_get_ems_atribut_FDD_TDD_new), -120))
            elif key_1 == "b2Threshold1GERAN" and value_1 != str(default_param.get("b2Threshold1GERAN").get(band_bw(dn, dict_get_ems_atribut_FDD_TDD_new), -120)):
                b2Threshold1GERAN_m120[key_1] = str(default_param.get("b2Threshold1GERAN").get(band_bw(dn, dict_get_ems_atribut_FDD_TDD_new), -120))
            elif key_1 == "arfcnValueListGERAN" and collections.Counter(value_1) != collections.Counter(value):
                slovar['arfcnValueListGERAN'] = value
            elif key_1 != "b2Threshold1GERANQci1" and key_1 != "b2Threshold1GERAN" and key_1 != "arfcnValueListGERAN" and value_1 != str(
                    default_param.get(key_1)):
                nccperm_255[key_1] = str(default_param.get(key_1))

            update1.update(slovar)
            update1.update(nccperm_255)
            update1.update(b2Threshold1GERAN_m120)
            update1.update(b2Threshold1GERANQci1_m120)
            if update1:
                update[dn] = update1
    print("update", update)

    dn_chahnge = dict_dn_cell_list_bcch.keys() - all_dn_lnhog_current
    list_dn_chahnge = list(dn_chahnge)
    for a in list_dn_chahnge:
        default_param_create = {}
        dict_valeu_def_param_change = {}
        for key_2param, value_2param in default_param.items():
            if key_2param == "b2Threshold1GERANQci1":
                dict_valeu_def_param_change[key_2param] = value_2param.get(band_bw(a, dict_get_ems_atribut_FDD_TDD_new), -120)
            if key_2param == "b2Threshold1GERAN":
                dict_valeu_def_param_change[key_2param] = value_2param.get(band_bw(a, dict_get_ems_atribut_FDD_TDD_new), -120)
        arfcnValueListGERAN = {}
        arfcnValueListGERAN["arfcnValueListGERAN"] = dict_dn_cell_list_bcch.get(a, dict_get_ems_atribut_FDD_TDD_new)
        default_param_create.update(default_param)
        default_param_create.update(arfcnValueListGERAN)
        default_param_create.update(dict_valeu_def_param_change)
        create[a] = default_param_create
    print("creat_sp", create)

        # delete
    for y in all_dn_lnhog_current:
        if y[-1] != "0":
            delete.append(y)
    print("delete", delete)


    # PUSH
    all_changes = len(create) + len(update) + len(delete)

    flag_alarm = 0
    if int(all_changes) > 0:
        print("{} differences found".format(int(all_changes)))
        if SON_MODE == "Closed Loop":
            print("Send changes to push")
            if maximum_changes > int(all_changes):
                try:
                    send_to_net = pusher(ems_client, region_name, create, update, delete)
                    print("Result push = ", send_to_net)
                except Exception as e:
                    print("Error in process push = ", e)
            else:
                flag_alarm = 1
                print("The number of changes exceeds the threshold of {} changes".format(maximum_changes))
        else:
            print("The module was launched in open loop")
    else:
        print("No changes to push")

    # report
    start_time = time.strftime("%d-%m-%Y_%H-%M-%S")
    stopt_time = time.strftime("%d-%m-%Y_%H-%M-%S")
    read_set_date = ReadSetData(script_data)
    get_class = CheckingChanges()
    rep = ReportGenerator(outputfile, start_time, stopt_time)

    name_list = ["Create", "Update", "Delete"]

    for i in name_list:
        curr_list = rep.create_new_sheets(i, ["Object", "Parameter", "Value", "Result"])




    if delete:
        real_val_delete = read_set_date.read_date(delete)
        delete_mo, not_delete_mo = get_class.check_delete(real_val_delete, delete)

        curr_list.UseSheet("Delete")

        if delete_mo:
            same_create_adapt = rep.data_transformation(delete_mo)
            curr_list.AddData(same_create_adapt)

        if not_delete_mo:
            diff_create_adapt = rep.data_transformation(not_delete_mo)
            curr_list.AddData(diff_create_adapt)

    if create:
        real_val_create = read_set_date.read_date(create)
        same_create, diff_create, not_exist_create = get_class.check_create_and_update(real_val_create, create)

        curr_list.UseSheet("Create")

        if same_create:
            same_create_adapt = rep.data_transformation(same_create)
            curr_list.AddData(same_create_adapt)

        if diff_create:
            diff_create_adapt = rep.data_transformation(diff_create)
            curr_list.AddData(diff_create_adapt)

        if not_exist_create:
            not_exist_create_adapt = rep.data_transformation(not_exist_create)
            curr_list.AddData(not_exist_create_adapt)

    if update:
        real_val_update = read_set_date.read_date(update)
        same_update, diff_update, not_exist_update = get_class.check_create_and_update(real_val_update, update)

        curr_list.UseSheet("Update")

        if same_update:
            same_create_adapt = rep.data_transformation(same_update)
            curr_list.AddData(same_create_adapt)

        if diff_update:
            diff_create_adapt = rep.data_transformation(diff_update)
            curr_list.AddData(diff_create_adapt)

        if not_exist_update:
            not_exist_create_adapt = rep.data_transformation(not_exist_update)
            curr_list.AddData(not_exist_create_adapt)


    rep.saved_file()

    # send email

    if (email_str and email_str.strip()) or (email_str_alarm and email_str_alarm.strip()):
        receivers = [r.strip() for r in re.split(';|,', email_str)]
        receivers_alarm = [r.strip() for r in re.split(';|,', email_str_alarm)]
        if receivers or receivers_alarm:
            emailer = script_data.get_emailer()
            if outputfile:
                attachments = [outputfile]
            else:
                attachments = None
            body = "*******THIS IS AN AUTO GENERATED EMAIL - PLEASE DO NOT REPLY********"
            try:
                # emailer.SendMail(receivers,
                #                  region_name + ' ' + EMAIL_SUBJECT + module_name + ' ' + config_report.Report_prefix,
                #                  body, attachments=attachments)
                if flag_alarm == 1:
                    emailer.SendMail(receivers_alarm, EMAIL_SUBJECT_ALARM + ' ' + region_name + ' all of changes: ' + str(all_changes) + ' module:' + module_name + ' ' + config_report.Report_prefix, body, attachments=attachments)
            except Exception:
                print("Email Sending Failed")
    else:
        print("!" * 100)
        print("No email address")
        print("!" * 100)



def pusher(ems_client, region_name, creat, update, delete):

    all_changes = len(creat) + len(update) + len(delete)
    print("{} differences found".format(all_changes))
    print("Number of create = ", len(creat))
    print("Number of update = ", len(update))
    print("Number of delete = ", len(delete))
    print("Send changes to push")

    result_push = ems_client.push_ems_attributes_by_oss(region_name, updates=update, creates=creat, deletes=delete)
    print("result_push = ", result_push)

    return result_push


def parser_exel(wb):
    name_sheets = wb.sheetnames
    for current_sheets in name_sheets:
        sheet = wb[current_sheets]
        row_M = sheet.max_row
        col_M = sheet.max_column
        parser_table = []
        slovar = {}
        if col_M > 1 or row_M > 1:
            for n in range(1, row_M + 1):
                param = sheet.cell(row=n, column=1).value
                for i in range(2, col_M + 1):
                    val = sheet.cell(row=n, column=i).value
                    parser_table.append(val)
                    slovar[param] = val
    slovar.pop('Parametr')
    return slovar

def band_bw(dn_lnhog_chek, dict_get_ems_atribut_FDD_TDD_new):
    """
    :param dn_lnhog:
    :return: band_bandwidth e.c "LTE_1800_20"
    """
    value_param = dict_get_ems_atribut_FDD_TDD_new.get(dn_lnhog_chek)
    earfcn = value_param.get("earfcnDL")
    if earfcn ==  None:
        earfcn = value_param.get("earfcn")
    band = def_band(int(earfcn))

    chBw = value_param.get("dlChBw")
    if chBw == None:
        chBw = value_param.get("chBw")
    bw = chBw.split(" ", 1)[0]
    band_bandwidth = band + "_" + str(bw)
    return band_bandwidth



def def_band(channel):
    """
    Detects LTE band by channel number
    :param channel: cell_from_dn.earfcn
    :return: e.g. "LTE_1800"
    """
    # print "check band"
    if 9870 <= channel <= 9919:
        band_f = "LTE_450"
    elif 6150 <= channel <= 6449:
        band_f = "LTE_800"
    elif 1200 <= channel <= 1949:
        band_f = "LTE_1800"
    elif 2750 <= channel <= 3449:
        band_f = "LTE_2600"
    elif 0 <= channel <= 599:
        band_f = "LTE_2100"
    elif 38650 <= channel <= 39649:
        band_f = "LTE_2300"
    elif 38250 <= channel <= 38649:
        band_f = "LTE_1900"
    elif 46709 <= channel <= 54539:
        band_f = "LTE_Unlicensed"
    else:
        band_f = "Unknown"
    return band_f


def GetConfigCategoryList():
    """Returns a list of categories that this module may use.
      If empty or the function is not present, only
      "Module Global Configuration" category is available. """
    return ["Custom_Configuration", ]




def GetEventTypes():
    """Return a list of trigger types"""
    return []


def GetDesc():
    """Return the description of this module"""
    return "This module determines the cosite of the umts sector, and enables / disables transition triggers"


def GetVersion():
    """Return the current version of this module"""
    return __version__


def GetParameters():
    """Return the list of parameters accepted by this module"""
    return Enet.ScriptParametersFromTuples(USER_PARAMETERS)


def GetScopeRules():
    """ Scope rules

    :return: the allowable vendors and technologies to run
    """
    return dict(tech=['LTE'],
                multi_tech=False,
                vendor=['nokia'],
                closed_loop=True)

